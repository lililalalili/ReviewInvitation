
r"""
Review Invitation Agent for Windows + Ubuntu-hosted Ollama
===============================================================

Real files detected:
- ReviewInvitationLog.xlsx   -> Sheet1 with headers: From | To
- NB_Author_2026.xlsm        -> Sheet 'Author'; if O/P/Q are missing, they are auto-created

Behavior:
1) Read the last row of ReviewInvitationLog.xlsx / Sheet1, column B ("To")
2) start_date = last To + 1 day
3) end_date = today
4) Query PubMed by Date - Publication [dp]-equivalent range for:
   - Nature
   - Science
   - Cell
   - Nature Neuroscience
   - Neuron
5) Save raw XML to D:\Agents\ReviewInvtation\PubMed\
6) Parse each article
7) Skip articles without abstract
8) For Nature / Science / Cell: send title + abstract to the configured LLM provider to judge neuroscience relevance
   For Neuron / Nature Neuroscience: treat as neuroscience by default
   Certain edge cases (retina, photoreceptor, BBB/CSF/choroid plexus, neurovascular,
   sensory epithelia, brain organoid, neural stem-like) are forced to neuroscience
9) Suggest concise invited review topic, field hierarchy, and author/email checks
9) Append accepted rows to NB_Author_2026.xlsm / Author
10) Append [From, To] to ReviewInvitationLog.xlsx / Sheet1 only after successful completion
11) Save detailed TSV run log under D:\Agents\ReviewInvtation\RunLogs\

Notes:
- Column 6 in the real workbook is "Date of Publication", so this script writes publication date there.
- Column 14 is "Date of Invitaion", so this script writes today's run date there.
- A SQLite state database is used to reduce duplicate imports and improve state reliability.
- The LLM runs on Ubuntu via a remote Ollama HTTP endpoint configured by OLLAMA_BASE_URL.
- Exact abstract/prompt/LLM response payloads are saved under RunLogs\\llm_audit\\YYYYMMDD\\ for verification.
- Macro Research Field is now forced to one fixed label from NEURO_MACRO_FIELDS.
"""

from __future__ import annotations

import json
import os
import re
import time
import hashlib
import sqlite3
from contextlib import contextmanager
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote_plus
import xml.etree.ElementTree as ET

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


# ------------------------------------------------------------------
# Paths
# ------------------------------------------------------------------

BASE_DIR = Path(os.getenv("REVIEW_INVITATION_BASE_DIR", r"D:\Agents\ReviewInvtation"))
LOG_XLSX = BASE_DIR / "ReviewInvitationLog.xlsx"
LOG_SHEET = "Sheet1"

TARGET_XLSM = BASE_DIR / "NB_Author_2026.xlsm"
TARGET_SHEET = "Author"

PUBMED_DIR = BASE_DIR / "PubMed"
RUN_LOG_DIR = BASE_DIR / "RunLogs"
STATE_DB = RUN_LOG_DIR / "review_invitation_agent_state.sqlite3"
RUN_LOCK_FILE = BASE_DIR / "review_invitation_agent.lock"
RUN_LOCK_STALE_HOURS = int(os.getenv("RUN_LOCK_STALE_HOURS", "12"))


# ------------------------------------------------------------------
# Search / model config
# ------------------------------------------------------------------

JOURNALS = ["Nature", "Science", "Cell", "Nature Neuroscience", "Neuron"]
REQUIRE_ABSTRACT = True
EXCLUDE_PT = ["erratum", "editorial", "comment", "review"]

AUTHOR_EXTRA_HEADERS = {
    15: "Macro Research Field",
    16: "Meso Research Field",
    17: "Micro Research Field",
    18: "Title (CN)",
    19: "Review Extension Potential",
    20: "Invited Review Angle",
    21: "Angle Rationale (CN)",
    22: "Editorial Bucket",
    23: "Manual Decision",
    24: "Neuro Decision Source",
    25: "PMID",
}

ALWAYS_NEURO_JOURNALS = {"neuron", "nature neuroscience"}
LLM_NEURO_JOURNALS = {"nature", "science", "cell"}

BROAD_SCOPE_NEURO_FIELDS = [
    "brain and central nervous system",
    "spinal cord and peripheral nervous system",
    "neurons, glia, synapses, and neural circuits",
    "neurodevelopment, neurogenesis, and neural stem cells",
    "neurodegeneration, neurological disease, and psychiatric disease with neural mechanisms",
    "retina, photoreceptors, and visual neuroscience",
    "blood-brain barrier, cerebrospinal fluid, choroid plexus, and neurovascular biology",
    "sensory epithelia and sensory transduction when the biological focus is neural or neuro-sensory",
    "brain organoids, neural stem-like systems, and neural differentiation models",
    "pain, itch, sensory processing, motor systems, cognition, emotion, sleep, and behavior with clear neural mechanisms",
    "neuroimmunology, neuroinflammation, neurometabolism, and mitochondrial neuroscience",
    "neurotechnology, neuroimaging, brain-computer interface, and neuroscience methods",
]

BROAD_SCOPE_NEURO_FIELDS_TEXT = "\n".join([f"- {x}" for x in BROAD_SCOPE_NEURO_FIELDS])

FORCED_NEURO_PATTERNS = [
    r"\bbrain\b",
    r"\bcns\b",
    r"\bcentral nervous system\b",
    r"\bspinal cord\b",
    r"\bperipheral nervous system\b",
    r"\bneural\b",
    r"\bneuron(?:s|al)?\b",
    r"\bglia(?:l)?\b",
    r"\bastrocyte(?:s)?\b",
    r"\bmicroglia(?:l)?\b",
    r"\boligodendrocyte(?:s)?\b",
    r"\bsynaps(?:e|es|tic)\b",
    r"\bconnectom(?:e|ics)\b",
    r"\bneurodevelopment(?:al)?\b",
    r"\bneurogenesis\b",
    r"\bneural stem(?: cell)?(?:s)?\b",
    r"\bbrain organoid(?:s)?\b",
    r"\bneural stem[- ]like\b",
    r"\bretina(?:l)?\b",
    r"\bphotoreceptor(?:s)?\b",
    r"\boptic nerve\b",
    r"\bblood[- ]brain barrier\b",
    r"\bbbb\b",
    r"\bcsf\b",
    r"\bcerebrospinal fluid\b",
    r"\bchoroid plexus\b",
    r"\bneurovascular\b",
    r"\bneurovascular unit\b",
    r"\bsensory epithel(?:ium|ia)\b",
    r"\bmechanosens(?:e|ation)\b",
    r"\bnocicept(?:ion|or)\b",
    r"\bpain\b",
    r"\bitch\b",
    r"\bhippocamp(?:us|al)\b",
    r"\bcortex|cortical\b",
    r"\bcerebell(?:um|ar)\b",
    r"\bthalam(?:us|ic)\b",
    r"\bhypothalam(?:us|ic)\b",
    r"\bamygdal(?:a|ar)\b",
    r"\baxon(?:al)?\b",
    r"\bdendrit(?:e|ic)\b",
    r"\bmyelin(?:ation)?\b",
    r"\bneurodegenerat(?:ion|ive)\b",
    r"\balzheimer'?s\b",
    r"\bparkinson'?s\b",
    r"\bals\b",
    r"\bepilep(?:sy|tic)\b",
    r"\bsleep\b",
    r"\bcircadian\b",
    r"\bcognit(?:ion|ive)\b",
    r"\bmemory\b",
    r"\bemotion(?:al)?\b",
    r"\bpsychiatric\b",
    r"\bneuroimmun(?:ology|e)\b",
    r"\bneuroinflammat(?:ion|ory)\b",
    r"\bneurometabol(?:ism|ic)\b",
    r"\bmitochondrial neuroscience\b",
    r"\bbrain[- ]computer interface\b",
    r"\bneuroimaging\b",
]


ANTI_NEURO_PATTERNS = [
    r"\bcardiac fibrosis\b",
    r"\bcardiomyocyte(?:s)?\b",
    r"\bheart failure\b",
    r"\bmyocard(?:ial|ium)\b",
    r"\bhepatocyte(?:s)?\b",
    r"\bliver\b",
    r"\bkidney\b",
    r"\brenal\b",
    r"\bpancrea(?:s|tic)\b",
    r"\bintestinal\b",
    r"\bgut\b",
    r"\bskeletal muscle\b",
    r"\badipose\b",
    r"\btumou?r microenvironment\b",
    r"\bnon[- ]small cell lung cancer\b",
    r"\bleukemia\b",
    r"\bmelanoma\b",
    r"\bcardiovascular\b",
    r"\bimmunotherapy\b",
    r"\bmetastasis\b",
    r"\bfibrosis\b",
]

RED_FLUORESCENT_FILL = PatternFill(fill_type="solid", fgColor="FF5C5C")
LIGHT_BLUE_ROW_FILL = PatternFill(fill_type="solid", fgColor="DDEEFF")
ORANGE_FONT_COLOR = "FF8C00"

PUBMED_TOOL = "NB_ReviewInvitationAgent"
PUBMED_EMAIL = os.getenv("PUBMED_EMAIL", "replace_with_your_email@example.com").strip()
PUBMED_API_KEY = (os.getenv("PUBMED_API_KEY") or "").strip() or None

REQUEST_CONNECT_TIMEOUT = int(os.getenv("REQUEST_CONNECT_TIMEOUT", "20"))
REQUEST_READ_TIMEOUT = int(os.getenv("REQUEST_READ_TIMEOUT", "180"))
REQUEST_RETRIES = int(os.getenv("REQUEST_RETRIES", "3"))
REQUEST_BACKOFF_FACTOR = float(os.getenv("REQUEST_BACKOFF_FACTOR", "1.5"))
REQUEST_PAUSE_SECONDS = 0.34

# LLM provider: "kimi" or "ollama"
LLM_PROVIDER = os.getenv("LLM_PROVIDER", "ollama").strip().lower()

# Kimi / Moonshot config
KIMI_BASE_URL = os.getenv("KIMI_BASE_URL", "https://api.moonshot.cn/v1").rstrip("/")
KIMI_API_KEY = (os.getenv("KIMI_API_KEY") or os.getenv("MOONSHOT_API_KEY") or "").strip()
KIMI_MODEL = os.getenv("KIMI_MODEL", "kimi-k2.5").strip()
KIMI_TIMEOUT_SECONDS = int(os.getenv("KIMI_TIMEOUT_SECONDS", "240"))
KIMI_DISABLE_THINKING = os.getenv("KIMI_DISABLE_THINKING", "true").strip().lower() in {"1", "true", "yes", "y"}

# Ubuntu-hosted Ollama (remote or tunneled)
# Examples:
#   http://192.168.1.50:11434/api
#   http://ubuntu-hostname:11434/api
#   http://127.0.0.1:11434/api   (if you use an SSH tunnel from Windows to Ubuntu)
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "").rstrip("/")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "nemotron-3-super:120b").strip()
OLLAMA_TIMEOUT_SECONDS = int(os.getenv("OLLAMA_TIMEOUT_SECONDS", "240"))
OLLAMA_KEEP_ALIVE = os.getenv("OLLAMA_KEEP_ALIVE", "30m").strip()
OLLAMA_BEARER_TOKEN = os.getenv("OLLAMA_BEARER_TOKEN", "").strip()

# LLM audit / debugging
LLM_AUDIT_ENABLED = os.getenv("LLM_AUDIT_ENABLED", "true").strip().lower() in {"1", "true", "yes", "y"}
LLM_AUDIT_SAVE_RESPONSE = os.getenv("LLM_AUDIT_SAVE_RESPONSE", "true").strip().lower() in {"1", "true", "yes", "y"}
PRINT_ABSTRACT_LENGTH = os.getenv("PRINT_ABSTRACT_LENGTH", "true").strip().lower() in {"1", "true", "yes", "y"}
ABSTRACT_SHORT_WARNING = int(os.getenv("ABSTRACT_SHORT_WARNING", "300"))
ABSTRACT_MIN_LENGTH = int(os.getenv("ABSTRACT_MIN_LENGTH", "200"))


NEURO_MACRO_FIELDS = [
    "Neurodevelopment & Neurogenesis",
    "Neural Stem Cells & Regeneration",
    "Neuronal Cell Biology",
    "Glia & Myelination",
    "Synapses & Neurotransmission",
    "Ion Channels & Excitability",
    "Neural Circuits & Connectomics",
    "Systems Neuroscience",
    "Sensory Neuroscience",
    "Motor Systems & Movement",
    "Autonomic & Neuroendocrine Neuroscience",
    "Learning, Memory & Plasticity",
    "Cognition & Executive Function",
    "Emotion, Motivation & Social Neuroscience",
    "Sleep & Circadian Neuroscience",
    "Pain, Itch & Interoception",
    "Neuroimmunology & Neuroinflammation",
    "Neurovascular Biology, BBB & CSF",
    "Neurometabolism & Mitochondrial Neuroscience",
    "Neurodegeneration & Aging",
    "Neurological Disorders",
    "Psychiatric Disorders",
    "Epilepsy & Network Excitability Disorders",
    "Brain Injury, Stroke & Repair",
    "Peripheral Nervous System & Neuromuscular Biology",
    "Computational & Theoretical Neuroscience",
    "Neurotechnology, Imaging & Methods",
    "Translational & Therapeutic Neuroscience",
]

NEURO_MACRO_FIELDS_TEXT = "\n".join([f"- {x}" for x in NEURO_MACRO_FIELDS])




def validate_runtime_config() -> None:
    if LLM_PROVIDER not in {"kimi", "ollama"}:
        raise RuntimeError(f"Unsupported LLM_PROVIDER: {LLM_PROVIDER}. Use 'kimi' or 'ollama'.")

    if LLM_PROVIDER == "ollama" and not OLLAMA_BASE_URL:
        raise RuntimeError(
            "OLLAMA_BASE_URL is not set. Set it to your Ubuntu Ollama host, for example "
            "'http://192.168.1.50:11434/api' or use an SSH tunnel and set "
            "'http://127.0.0.1:11434/api'."
        )

    if LLM_PROVIDER == "kimi" and not KIMI_API_KEY:
        raise RuntimeError(
            "KIMI_API_KEY is not set. Please set environment variable KIMI_API_KEY "
            "(or MOONSHOT_API_KEY) before running this script."
        )

    if not PUBMED_EMAIL or PUBMED_EMAIL == "replace_with_your_email@example.com":
        print("WARNING: PUBMED_EMAIL is not set. NCBI recommends setting a real email address.")

# ------------------------------------------------------------------
# Country normalization
# ------------------------------------------------------------------

CHINA_EQUIVALENTS = {
    "china",
    "people's republic of china",
    "p.r. china",
    "pr china",
    "mainland china",
    "hong kong, china",
}


# ------------------------------------------------------------------
# Data structures
# ------------------------------------------------------------------


def normalize_review_extension_potential(value: str) -> str:
    v = clean_space(value).lower()
    if v == "high":
        return "High"
    if v == "medium":
        return "Medium"
    if v == "low":
        return "Low"
    return "Medium"


def infer_neuro_decision_source(reason: str, journal: str) -> str:
    reason_l = clean_space(reason).lower()
    journal_l = clean_space(journal).lower()
    if journal_l in ALWAYS_NEURO_JOURNALS or "journal policy auto-accept" in reason_l or "always-neuro journal rule" in reason_l:
        return "Journal Rule"
    if "forced neuroscience topic override" in reason_l or "broad-scope neuroscience rule" in reason_l:
        return "Broad Rule"
    return "LLM"


def acquire_run_lock(lock_path: Path) -> None:
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    if lock_path.exists():
        try:
            age_hours = (time.time() - lock_path.stat().st_mtime) / 3600.0
        except Exception:
            age_hours = 0
        if age_hours < RUN_LOCK_STALE_HOURS:
            raise RuntimeError(
                f"Another run appears active because lock file exists: {lock_path}. "
                f"If this is stale, delete it manually or wait until it is older than {RUN_LOCK_STALE_HOURS} hours."
            )
        try:
            lock_path.unlink()
        except Exception:
            pass
    lock_path.write_text(
        json.dumps({"pid": os.getpid(), "time": datetime.now().isoformat()}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def release_run_lock(lock_path: Path) -> None:
    try:
        if lock_path.exists():
            lock_path.unlink()
    except Exception:
        pass


@dataclass
class ParsedArticle:
    pmid: str
    journal: str
    title: str
    abstract: str

    publication_date: Optional[date]

    first_author_full_name: str
    first_author_email: str

    last_author_family_name: str
    last_author_full_name: str
    last_author_email: str
    last_author_affiliation: str
    last_author_country: str

    pubmed_link: str
    neuro_decision_source: str = ""


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def ensure_dirs() -> None:
    PUBMED_DIR.mkdir(parents=True, exist_ok=True)
    RUN_LOG_DIR.mkdir(parents=True, exist_ok=True)
    init_state_db()


def fmt_slash(d: date) -> str:
    return d.strftime("%Y/%m/%d")


def fmt_compact(d: date) -> str:
    return d.strftime("%Y%m%d")


def clean_space(text: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def extract_email(text: str) -> str:
    if not text:
        return ""
    m = re.search(r'([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})', text)
    return m.group(1) if m else ""


def normalize_review_extension_potential(value: str) -> str:
    v = clean_space(value).lower()
    if v == "high":
        return "High"
    if v == "medium":
        return "Medium"
    if v == "low":
        return "Low"
    return "Medium"


def choose_best_affiliation_and_email(author_el: ET.Element) -> Tuple[str, str]:
    affs = author_affiliations(author_el)
    if not affs:
        return "", ""

    # Prefer explicit electronic-address affiliations
    for aff in affs:
        if re.search(r"electronic address\s*:", aff, flags=re.I):
            email = extract_email(aff)
            return aff, email

    # Then prefer any affiliation containing an email
    for aff in affs:
        email = extract_email(aff)
        if email:
            return aff, email

    # Fallback to first affiliation
    return affs[0], ""


def init_state_db() -> None:
    RUN_LOG_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(STATE_DB)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS processed_articles (
                pmid TEXT PRIMARY KEY,
                first_seen_at TEXT,
                last_seen_at TEXT,
                status TEXT
            )
            """
        )
        conn.commit()
    finally:
        conn.close()




def contains_forced_neuro_topic(title: str, abstract: str) -> bool:
    hay = f"{title}\n{abstract}".lower()
    return any(re.search(p, hay, flags=re.I) for p in FORCED_NEURO_PATTERNS)


def orange_font_like(font_obj):
    base = copy(font_obj) if font_obj else Font()
    base.color = ORANGE_FONT_COLOR
    return base


def classify_neuro_by_policy(rec) -> Tuple[str, float, str, str]:
    journal_key = clean_space(rec.journal).lower()
    if journal_key in ALWAYS_NEURO_JOURNALS:
        return "Yes", 1.0, f"Journal policy auto-accept: {rec.journal}", "Journal Rule"
    if contains_forced_neuro_topic(rec.title, rec.abstract):
        return "Yes", 1.0, "Forced neuroscience topic override", "Broad Rule"
    label, conf, reason = classify_neuroscience(rec.pmid, rec.title, rec.abstract)
    return label, conf, reason, "LLM"


def safe_file_component(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", (text or "").strip())[:120] or "unknown"


def sha256_text(text: str) -> str:
    return hashlib.sha256((text or "").encode("utf-8")).hexdigest()


def llm_audit_dir(run_day: Optional[date] = None) -> Path:
    day = run_day or date.today()
    out = RUN_LOG_DIR / "llm_audit" / fmt_compact(day)
    out.mkdir(parents=True, exist_ok=True)
    return out


def save_llm_request_audit(
    *,
    pmid: str,
    stage: str,
    provider: str,
    title: str,
    abstract: str,
    prompt: str,
    schema: dict,
    request_payload: dict,
) -> Path:
    base = llm_audit_dir() / f"{safe_file_component(pmid)}_{safe_file_component(stage)}"
    payload = {
        "pmid": pmid,
        "stage": stage,
        "provider": provider,
        "title": title,
        "title_len": len(title or ""),
        "title_sha256": sha256_text(title or ""),
        "abstract": abstract,
        "abstract_len": len(abstract or ""),
        "abstract_sha256": sha256_text(abstract or ""),
        "prompt": prompt,
        "prompt_len": len(prompt or ""),
        "prompt_sha256": sha256_text(prompt or ""),
        "schema": schema,
        "request_payload": request_payload,
    }
    req_path = base.with_name(base.name + "_request.json")
    req_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return base


def save_llm_response_audit(
    *,
    base_path: Path,
    provider: str,
    http_response_json: Optional[dict],
    parsed_json: Optional[dict],
    raw_content: str,
    error: str = "",
) -> Path:
    payload = {
        "provider": provider,
        "raw_content": raw_content,
        "raw_content_len": len(raw_content or ""),
        "raw_content_sha256": sha256_text(raw_content or ""),
        "parsed_json": parsed_json,
        "http_response_json": http_response_json,
        "error": error,
    }
    resp_path = base_path.with_name(base_path.name + "_response.json")
    resp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return resp_path


def normalize_country(country: str) -> str:
    c = clean_space(country).rstrip(".")
    low = c.lower()
    if low in {"usa", "u.s.a.", "u.s.a", "u.s.", "united states", "united states of america"}:
        return "USA"
    if low in {"uk", "u.k.", "united kingdom", "england"}:
        return "UK"
    if low in CHINA_EQUIVALENTS:
        return "China"
    return c


def infer_country_from_affiliation(aff: str) -> str:
    if not aff:
        return ""

    # remove electronic address tail
    aff_wo_email = re.sub(r"Electronic address:.*$", "", aff, flags=re.I).strip().rstrip(".")
    if not aff_wo_email:
        return ""

    # split on semicolons first, keep last non-empty block
    parts = [p.strip() for p in aff_wo_email.split(";") if p.strip()]
    block = parts[-1] if parts else aff_wo_email

    comma_parts = [p.strip() for p in block.split(",") if p.strip()]
    if comma_parts:
        return normalize_country(comma_parts[-1])

    return normalize_country(block)


def parse_pubmed_month(month_text: str) -> int:
    month_text = clean_space(month_text)
    if not month_text:
        return 1

    month_map = {
        "Jan": 1, "January": 1,
        "Feb": 2, "February": 2,
        "Mar": 3, "March": 3,
        "Apr": 4, "April": 4,
        "May": 5,
        "Jun": 6, "June": 6,
        "Jul": 7, "July": 7,
        "Aug": 8, "August": 8,
        "Sep": 9, "Sept": 9, "September": 9,
        "Oct": 10, "October": 10,
        "Nov": 11, "November": 11,
        "Dec": 12, "December": 12,
    }
    return month_map.get(month_text, 1)


def parse_article_date(article_el: ET.Element) -> Optional[date]:
    # Prefer explicit ArticleDate if present
    article_date = article_el.find("./ArticleDate")
    if article_date is not None:
        y = clean_space(article_date.findtext("Year"))
        m = clean_space(article_date.findtext("Month"))
        d = clean_space(article_date.findtext("Day"))
        if y and m and d:
            try:
                return date(int(y), int(m), int(d))
            except Exception:
                pass

    pubdate = article_el.find("./Journal/JournalIssue/PubDate")
    if pubdate is not None:
        y = clean_space(pubdate.findtext("Year"))
        medline_date = clean_space(pubdate.findtext("MedlineDate"))
        m = clean_space(pubdate.findtext("Month"))
        d = clean_space(pubdate.findtext("Day")) or "1"

        if y:
            try:
                month_num = parse_pubmed_month(m) if m else 1
                return date(int(y), month_num, int(d))
            except Exception:
                pass

        # crude MedlineDate fallback, e.g. "2026 Jan 29"
        if medline_date:
            m2 = re.search(r"(\d{4})\s+([A-Za-z]+)\s+(\d{1,2})", medline_date)
            if m2:
                try:
                    return date(int(m2.group(1)), parse_pubmed_month(m2.group(2)), int(m2.group(3)))
                except Exception:
                    pass
            m3 = re.search(r"(\d{4})\s+([A-Za-z]+)", medline_date)
            if m3:
                try:
                    return date(int(m3.group(1)), parse_pubmed_month(m3.group(2)), 1)
                except Exception:
                    pass
            m4 = re.search(r"(\d{4})", medline_date)
            if m4:
                try:
                    return date(int(m4.group(1)), 1, 1)
                except Exception:
                    pass

    return None


def article_title(article_el: ET.Element) -> str:
    t = article_el.find("ArticleTitle")
    return clean_space("".join(t.itertext())) if t is not None else ""


def abstract_text(article_el: ET.Element) -> str:
    nodes = article_el.findall("./Abstract/AbstractText")
    if not nodes:
        nodes = article_el.findall(".//OtherAbstract/AbstractText")

    parts = []
    for ab in nodes:
        label = clean_space(ab.attrib.get("Label") or ab.attrib.get("NlmCategory"))
        txt = clean_space("".join(ab.itertext()))
        if txt:
            parts.append(f"{label}: {txt}" if label else txt)

    return "\n\n".join(parts).strip()


def author_full_name(author_el: ET.Element) -> str:
    collective = clean_space(author_el.findtext("CollectiveName"))
    if collective:
        return collective
    last = clean_space(author_el.findtext("LastName"))
    fore = clean_space(author_el.findtext("ForeName"))
    if last and fore:
        return f"{fore} {last}"
    return fore or last


def author_family_name(author_el: ET.Element) -> str:
    return clean_space(author_el.findtext("LastName"))


def author_affiliations(author_el: ET.Element) -> List[str]:
    out = []
    for aff in author_el.findall("./AffiliationInfo/Affiliation"):
        text = clean_space("".join(aff.itertext()))
        if text:
            out.append(text)
    return out


def primary_affiliation(author_el: ET.Element) -> str:
    affs = author_affiliations(author_el)
    return affs[0] if affs else ""


def load_processed_pmids() -> set[str]:
    init_state_db()
    conn = sqlite3.connect(STATE_DB)
    try:
        cur = conn.execute("SELECT pmid FROM processed_articles WHERE status = 'processed'")
        return {row[0] for row in cur.fetchall()}
    finally:
        conn.close()


def save_processed_pmids(pmids: set[str]) -> None:
    init_state_db()
    conn = sqlite3.connect(STATE_DB)
    now = datetime.now().isoformat()
    try:
        for pmid in sorted(pmids):
            conn.execute(
                """
                INSERT INTO processed_articles (pmid, first_seen_at, last_seen_at, status)
                VALUES (?, ?, ?, 'processed')
                ON CONFLICT(pmid) DO UPDATE SET
                    last_seen_at=excluded.last_seen_at,
                    status='processed'
                """,
                (pmid, now, now),
            )
        conn.commit()
    finally:
        conn.close()


# ------------------------------------------------------------------
# Workbook handling
# ------------------------------------------------------------------

def read_last_to_date(log_xlsx: Path, sheet_name: str) -> date:
    wb = load_workbook(log_xlsx, data_only=True)
    ws = wb[sheet_name]

    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=2).value
        if val:
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date):
                return val
            if isinstance(val, str):
                for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
                    try:
                        return datetime.strptime(val.strip(), fmt).date()
                    except Exception:
                        pass

    raise ValueError("Could not find the last date in column B ('To') of ReviewInvitationLog.xlsx")


def append_from_to_log(log_xlsx: Path, sheet_name: str, from_date: date, to_date: date) -> None:
    last_err = None
    for _ in range(3):
        try:
            wb = load_workbook(log_xlsx)
            ws = wb[sheet_name]
            row = ws.max_row + 1

            # copy style from previous data row if present
            src_row = max(2, row - 1)
            for col in (1, 2):
                src = ws.cell(src_row, col)
                dst = ws.cell(row, col)

                dst.value = from_date if col == 1 else to_date
                if src.has_style:
                    dst._style = copy(src._style)
                if src.number_format:
                    dst.number_format = src.number_format

            wb.save(log_xlsx)
            return
        except PermissionError as e:
            last_err = e
            time.sleep(2)

    raise RuntimeError(
        f"Could not update log file after retries: {log_xlsx}. "
        "Please close Excel or pause sync tools and try again."
    ) from last_err


def load_existing_author_keys(target_xlsm: Path, sheet_name: str) -> tuple[set[tuple[str, str, str]], set[str]]:
    wb = load_workbook(target_xlsm, data_only=True, keep_vba=True)
    ws = wb[sheet_name]
    keys = set()
    pmids = set()
    pmid_col = 25
    for row in range(2, ws.max_row + 1):
        email = clean_space(str(ws.cell(row, 2).value or ""))
        journal = clean_space(str(ws.cell(row, 3).value or ""))
        title = clean_space(str(ws.cell(row, 4).value or ""))
        pmid = clean_space(str(ws.cell(row, pmid_col).value or ""))
        if email or title:
            keys.add((email.lower(), journal.lower(), title.lower()))
        if pmid:
            pmids.add(pmid)
    return keys, pmids


# ------------------------------------------------------------------
# PubMed API
# ------------------------------------------------------------------

def build_pubmed_query(start_date: date, end_date: date) -> str:
    journal_clause = " OR ".join([f'"{j}"[ta]' for j in JOURNALS])
    exclusion = " OR ".join([f"{pt}[pt]" for pt in EXCLUDE_PT])
    q = (
        f"({journal_clause}) "
        f'AND ("{fmt_slash(start_date)}"[Date - Publication] : "{fmt_slash(end_date)}"[Date - Publication]) '
        f"NOT ({exclusion})"
    )
    if REQUIRE_ABSTRACT:
        q += " AND hasabstract"
    return q


def build_requests_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=REQUEST_RETRIES,
        connect=REQUEST_RETRIES,
        read=REQUEST_RETRIES,
        status=REQUEST_RETRIES,
        backoff_factor=REQUEST_BACKOFF_FACTOR,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET", "POST"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({
        "User-Agent": f"{PUBMED_TOOL}/1.0 ({PUBMED_EMAIL})",
        "Accept": "*/*",
    })
    return session


HTTP_SESSION = build_requests_session()


def http_get(url: str, *, params: Optional[Dict[str, str]] = None, timeout: Optional[Tuple[int, int]] = None, headers: Optional[Dict[str, str]] = None) -> requests.Response:
    timeout = timeout or (REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT)
    resp = HTTP_SESSION.get(url, params=params, timeout=timeout, headers=headers)
    resp.raise_for_status()
    return resp


def http_post(url: str, *, json_body: Optional[dict] = None, timeout: Optional[int] = None, headers: Optional[Dict[str, str]] = None) -> requests.Response:
    timeout = timeout or OLLAMA_TIMEOUT_SECONDS
    resp = HTTP_SESSION.post(url, json=json_body, timeout=timeout, headers=headers)
    resp.raise_for_status()
    return resp


def ncbi_get(url: str, params: Dict[str, str]) -> requests.Response:
    p = dict(params)
    p["tool"] = PUBMED_TOOL
    p["email"] = PUBMED_EMAIL
    if PUBMED_API_KEY:
        p["api_key"] = PUBMED_API_KEY

    resp = http_get(url, params=p)
    time.sleep(REQUEST_PAUSE_SECONDS)
    return resp


def esearch_pubmed(query: str) -> List[str]:
    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {
        "db": "pubmed",
        "term": query,
        "retmode": "json",
        "retmax": "10000",
        "sort": "pub_date",
    }
    resp = ncbi_get(url, params)
    data = resp.json()
    return data["esearchresult"].get("idlist", [])


def efetch_pubmed_xml(pmids: List[str]) -> str:
    if not pmids:
        return "<PubmedArticleSet></PubmedArticleSet>"

    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    chunks = [pmids[i:i + 200] for i in range(0, len(pmids), 200)]
    xml_parts = ["<PubmedArticleSet>"]

    for chunk in chunks:
        params = {"db": "pubmed", "id": ",".join(chunk), "retmode": "xml"}
        resp = ncbi_get(url, params)
        text = resp.text.strip()
        text = re.sub(r"^\s*<\?xml[^>]*\?>", "", text)
        text = re.sub(r"^\s*<!DOCTYPE[^>]*>", "", text)
        text = re.sub(r"^\s*<PubmedArticleSet>", "", text)
        text = re.sub(r"</PubmedArticleSet>\s*$", "", text)
        xml_parts.append(text)

    xml_parts.append("</PubmedArticleSet>")
    return "\n".join(xml_parts)


def save_xml(xml_text: str, end_date: date) -> Path:
    out = PUBMED_DIR / f"pubmed-CellNatureNeuron-set-{fmt_compact(end_date)}.xml"
    out.write_text(xml_text, encoding="utf-8")
    return out


# ------------------------------------------------------------------
# Parse XML
# ------------------------------------------------------------------

def parse_pubmed_xml(xml_text: str) -> List[ParsedArticle]:
    root = ET.fromstring(xml_text)
    rows: List[ParsedArticle] = []

    for rec in root.findall("./PubmedArticle"):
        pmid = clean_space(rec.findtext("./MedlineCitation/PMID"))
        article_el = rec.find("./MedlineCitation/Article")
        if article_el is None:
            continue

        authors = article_el.findall("./AuthorList/Author")
        if not authors:
            continue

        first_author = authors[0]
        last_author = authors[-1]

        first_name = author_full_name(first_author)
        last_name_full = author_full_name(last_author)
        last_family = author_family_name(last_author)

        first_aff, first_email = choose_best_affiliation_and_email(first_author)
        last_aff, last_email = choose_best_affiliation_and_email(last_author)
        country = infer_country_from_affiliation(last_aff)

        title = article_title(article_el)
        abstract = abstract_text(article_el)
        journal = clean_space(article_el.findtext("./Journal/Title"))
        pub_date = parse_article_date(article_el)

        pubmed_link = f"https://pubmed.ncbi.nlm.nih.gov/?term={quote_plus(last_email)}" if last_email else ""

        rows.append(
            ParsedArticle(
                pmid=pmid,
                journal=journal,
                title=title,
                abstract=abstract,
                publication_date=pub_date,
                first_author_full_name=first_name,
                first_author_email=first_email,
                last_author_family_name=last_family,
                last_author_full_name=last_name_full,
                last_author_email=last_email,
                last_author_affiliation=last_aff,
                last_author_country=country,
                pubmed_link=pubmed_link,
            )
        )

    return rows


# ------------------------------------------------------------------
# Ollama
# ------------------------------------------------------------------

CLASSIFY_SCHEMA = {
    "type": "object",
    "properties": {
        "is_neuroscience_related": {"type": "string", "enum": ["Yes", "No"]},
        "confidence": {"type": "number"},
        "reason": {"type": "string"},
    },
    "required": ["is_neuroscience_related", "confidence", "reason"],
}

TOPIC_SCHEMA = {
    "type": "object",
    "properties": {
        "invited_review_topic": {"type": "string"},
    },
    "required": ["invited_review_topic"],
}


FIELD_HIERARCHY_SCHEMA = {
    "type": "object",
    "properties": {
        "macro_research_field": {"type": "string", "enum": NEURO_MACRO_FIELDS},
        "meso_research_field": {"type": "string"},
        "micro_research_field": {"type": "string"},
    },
    "required": ["macro_research_field", "meso_research_field", "micro_research_field"],
}

TITLE_TRANSLATION_SCHEMA = {
    "type": "object",
    "properties": {
        "title_zh_cn": {"type": "string"},
    },
    "required": ["title_zh_cn"],
}

REVIEW_ANGLE_SCHEMA = {
    "type": "object",
    "properties": {
        "review_extension_potential": {"type": "string", "enum": ["High", "Medium", "Low"]},
        "invited_review_angle": {"type": "string"},
        "angle_rationale_zh_cn": {"type": "string"},
    },
    "required": ["review_extension_potential", "invited_review_angle", "angle_rationale_zh_cn"],
}


AUTHOR_IDENTITY_SCHEMA = {
    "type": "object",
    "properties": {
        "last_author_email_match": {"type": "string", "enum": ["Yes", "No", "Unknown"]},
        "first_author_email_match": {"type": "string", "enum": ["Yes", "No", "Unknown"]},
        "last_author_is_chinese_name": {"type": "string", "enum": ["Yes", "No", "Unclear"]},
        "first_author_is_chinese_name": {"type": "string", "enum": ["Yes", "No", "Unclear"]},
        "reason": {"type": "string"},
    },
    "required": [
        "last_author_email_match",
        "first_author_email_match",
        "last_author_is_chinese_name",
        "first_author_is_chinese_name",
        "reason",
    ],
}


def kimi_generate(prompt: str, schema: dict, *, audit_meta: Optional[dict] = None) -> dict:
    """
    Kimi chat-completions wrapper.
    Uses JSON mode, so the prompt must explicitly ask for JSON only.
    """
    url = f"{KIMI_BASE_URL}/chat/completions"
    payload = {
        "model": KIMI_MODEL,
        "messages": [
            {"role": "system", "content": "You are a precise scientific screening assistant. Output valid JSON only."},
            {"role": "user", "content": prompt},
        ],
        "response_format": {"type": "json_object"},
        "max_completion_tokens": 512,
        "stream": False,
    }
    if KIMI_DISABLE_THINKING and KIMI_MODEL == "kimi-k2.5":
        payload["thinking"] = {"type": "disabled"}

    audit_base = None
    if audit_meta and LLM_AUDIT_ENABLED:
        audit_base = save_llm_request_audit(
            pmid=str(audit_meta.get("pmid", "")),
            stage=str(audit_meta.get("stage", "")),
            provider="kimi",
            title=str(audit_meta.get("title", "")),
            abstract=str(audit_meta.get("abstract", "")),
            prompt=prompt,
            schema=schema,
            request_payload=payload,
        )

    headers = {
        "Authorization": f"Bearer {KIMI_API_KEY}",
        "Content-Type": "application/json",
    }
    resp = http_post(url, json_body=payload, headers=headers, timeout=KIMI_TIMEOUT_SECONDS)
    resp.raise_for_status()
    data = resp.json()
    raw = data["choices"][0]["message"]["content"]

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        if audit_base and LLM_AUDIT_SAVE_RESPONSE:
            save_llm_response_audit(
                base_path=audit_base,
                provider="kimi",
                http_response_json=data,
                parsed_json=None,
                raw_content=raw,
                error=f"JSONDecodeError: {e}",
            )
        raise ValueError(f"Kimi returned non-JSON content: {raw[:500]}") from e

    if audit_base and LLM_AUDIT_SAVE_RESPONSE:
        save_llm_response_audit(
            base_path=audit_base,
            provider="kimi",
            http_response_json=data,
            parsed_json=parsed,
            raw_content=raw,
            error="",
        )
    return parsed



def ollama_headers() -> dict:
    headers = {"Content-Type": "application/json"}
    if OLLAMA_BEARER_TOKEN:
        headers["Authorization"] = f"Bearer {OLLAMA_BEARER_TOKEN}"
    return headers


def ollama_root_base() -> str:
    if OLLAMA_BASE_URL.endswith("/api"):
        return OLLAMA_BASE_URL[:-4]
    return OLLAMA_BASE_URL


def verify_ollama_remote() -> List[str]:
    url = f"{ollama_root_base()}/api/tags"
    resp = http_get(url, headers=ollama_headers(), timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT))
    resp.raise_for_status()
    data = resp.json()
    models = [m.get("name", "") for m in data.get("models", []) if m.get("name")]
    return models


def ollama_generate(prompt: str, schema: dict, *, audit_meta: Optional[dict] = None) -> dict:
    url = f"{OLLAMA_BASE_URL}/generate"
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "format": schema,
        "keep_alive": OLLAMA_KEEP_ALIVE,
    }

    audit_base = None
    if audit_meta and LLM_AUDIT_ENABLED:
        audit_base = save_llm_request_audit(
            pmid=str(audit_meta.get("pmid", "")),
            stage=str(audit_meta.get("stage", "")),
            provider="ollama",
            title=str(audit_meta.get("title", "")),
            abstract=str(audit_meta.get("abstract", "")),
            prompt=prompt,
            schema=schema,
            request_payload=payload,
        )

    try:
        resp = http_post(url, json_body=payload, headers=ollama_headers(), timeout=OLLAMA_TIMEOUT_SECONDS)
    except requests.exceptions.RequestException as e:
        raise RuntimeError(
            "Could not reach the remote Ollama server on Ubuntu. "
            f"OLLAMA_BASE_URL={OLLAMA_BASE_URL}. "
            "Check that Ubuntu is reachable from Windows, Ollama is listening on 0.0.0.0:11434 or that your SSH tunnel is up, "
            "and that firewall rules allow the connection."
        ) from e
    resp.raise_for_status()
    data = resp.json()
    raw = data.get("response", "")

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        if audit_base and LLM_AUDIT_SAVE_RESPONSE:
            save_llm_response_audit(
                base_path=audit_base,
                provider="ollama",
                http_response_json=data,
                parsed_json=None,
                raw_content=raw,
                error=f"JSONDecodeError: {e}",
            )
        raise

    if audit_base and LLM_AUDIT_SAVE_RESPONSE:
        save_llm_response_audit(
            base_path=audit_base,
            provider="ollama",
            http_response_json=data,
            parsed_json=parsed,
            raw_content=raw,
            error="",
        )
    return parsed


def llm_generate(prompt: str, schema: dict, *, audit_meta: Optional[dict] = None) -> dict:
    if LLM_PROVIDER.lower() == "kimi":
        return kimi_generate(prompt, schema, audit_meta=audit_meta)
    return ollama_generate(prompt, schema, audit_meta=audit_meta)


def classify_neuroscience(pmid: str, title: str, abstract: str) -> Tuple[str, float, str]:
    prompt = f"""
You are screening articles for Neuroscience Bulletin invited review candidates.

Return JSON only.

Decision standard:
- YES: directly about the nervous system, brain, spinal cord, peripheral nerves, neurons, glia,
  synapses, neural circuits, neurodevelopment, neurodegeneration, neural injury, pain/itch,
  sensory neuroscience, neuroimaging, cognition, behavior with neural mechanisms,
  psychiatric/neurological disease with mechanistic neural focus, or neuroscience methods.
- YES also for the following domains: retina, photoreceptor biology, blood-brain barrier (BBB),
  cerebrospinal fluid (CSF), choroid plexus, neurovascular biology, sensory epithelia,
  brain organoids, and neural stem-like systems when the paper's central biological focus is neural.
- NO: mainly oncology, immunology, metabolism, general cell biology, cardiology, or other fields
  without a clear neural focus.
- A paper is NOT neuroscience merely because the nervous system is a peripheral context,
  secondary phenotype, or speculative downstream implication.

Title: {title}

Abstract:
{abstract}
""".strip()

    result = llm_generate(
        prompt,
        CLASSIFY_SCHEMA,
        audit_meta={
            "pmid": pmid,
            "stage": "classify",
            "title": title,
            "abstract": abstract,
        },
    )
    return (
        clean_space(result.get("is_neuroscience_related", "No")),
        float(result.get("confidence", 0)),
        clean_space(result.get("reason", "")),
    )


def is_forced_broad_scope_neuro(title: str, abstract: str) -> bool:
    text = f"{title}\n{abstract}".lower()
    return any(re.search(pattern, text, flags=re.I) for pattern in FORCED_NEURO_PATTERNS)

def is_anti_neuro_context(title: str, abstract: str) -> bool:
    text = f"{title}\n{abstract}".lower()
    return any(re.search(pattern, text, flags=re.I) for pattern in ANTI_NEURO_PATTERNS)

def decide_neuroscience_status(pmid: str, journal: str, title: str, abstract: str) -> Tuple[str, float, str, str]:
    journal_l = clean_space(journal).lower()
    if journal_l in ALWAYS_NEURO_JOURNALS:
        return "Yes", 1.00, "Always-neuro journal rule", "Journal Rule"

    broad_neuro = is_forced_broad_scope_neuro(title, abstract)
    anti_neuro = is_anti_neuro_context(title, abstract)

    if journal_l in LLM_NEURO_JOURNALS and broad_neuro and not anti_neuro:
        return "Yes", 0.99, "Broad-scope neuroscience rule match", "Broad Rule"

    label, conf, reason = classify_neuroscience(pmid, title, abstract)
    return label, conf, reason, "LLM"


def default_author_identity() -> dict:
    return {
        "last_author_email_match": "Unknown",
        "first_author_email_match": "Unknown",
        "last_author_is_chinese_name": "Unclear",
        "first_author_is_chinese_name": "Unclear",
        "reason": "",
    }


def assess_author_identity(
    pmid: str,
    last_author_name: str,
    last_author_email: str,
    first_author_name: str,
    first_author_email: str,
) -> dict:
    prompt = f"""
You are helping a scientific editor verify author-name/email consistency and infer whether a name is likely Chinese.

Return JSON only.

Tasks:
1) Decide whether the last author's name matches the last author's email.
2) Decide whether the first author's name matches the first author's email.
3) Decide whether the last author's name is likely a Chinese personal name.
4) Decide whether the first author's name is likely a Chinese personal name.

Rules:
- If an email is missing or empty, the corresponding email_match must be "Unknown".
- Use "Yes" when the email plausibly belongs to the person named.
- Use "No" when the email clearly appears to belong to someone else.
- Use "Unknown" only when email is missing or impossible to judge.
- For Chinese-name inference, use:
  - "Yes" if the name is likely Chinese,
  - "No" if likely not Chinese,
  - "Unclear" if genuinely uncertain.
- Base judgments only on the names and emails below.

Last author name: {last_author_name}
Last author email: {last_author_email}

First author name: {first_author_name}
First author email: {first_author_email}
""".strip()

    result = llm_generate(
        prompt,
        AUTHOR_IDENTITY_SCHEMA,
        audit_meta={
            "pmid": pmid,
            "stage": "author_identity",
            "title": f"{last_author_name} | {first_author_name}",
            "abstract": f"last_email={last_author_email}\nfirst_email={first_author_email}",
        },
    )

    merged = default_author_identity()
    for key in merged:
        if key in result:
            merged[key] = clean_space(str(result.get(key, merged[key])))
    if not clean_space(last_author_email):
        merged["last_author_email_match"] = "Unknown"
    if not clean_space(first_author_email):
        merged["first_author_email_match"] = "Unknown"
    return merged


def suggest_topic(pmid: str, title: str, abstract: str) -> str:
    prompt = f"""
You are helping a scientific editor invite a review article for Neuroscience Bulletin.

Return JSON only.
Propose one concise invited review direction based on this paper.

Rules:
- 3 to 10 words preferred
- specific, not generic
- must be neuroscience-relevant
- not a full sentence

Title: {title}

Abstract:
{abstract}
""".strip()

    result = llm_generate(
        prompt,
        TOPIC_SCHEMA,
        audit_meta={
            "pmid": pmid,
            "stage": "topic",
            "title": title,
            "abstract": abstract,
        },
    )
    return clean_space(result.get("invited_review_topic", ""))[:200]


def generate_review_angle_package(
    pmid: str,
    title: str,
    abstract: str,
    journal: str,
    macro_field: str,
    meso_field: str,
    micro_field: str,
) -> Tuple[str, str, str]:
    prompt = f"""
You are a senior scientific editor evaluating whether a neuroscience paper can be expanded into an invited review article for a neuroscience journal.

Return JSON only.

Your job is not to summarize the paper. Your job is to identify whether the paper supports a broader, literature-expandable review theme, and if so, define the best invitation angle.

Based on the journal, title, abstract, and field hierarchy, provide:
1) review_extension_potential: High / Medium / Low
2) invited_review_angle: one concise invited-review angle in English
3) angle_rationale_zh_cn: one concise Chinese sentence explaining why this angle is suitable

Core editorial standard:
- A good invited review angle should be broader than the focal paper but still specific enough to organize a review.
- It should connect the focal paper to a wider body of literature, not just restate the paper.
- It should sound like a plausible review title, invitation theme, or central review heading for a neuroscience journal.
- Prefer themes based on mechanisms, biological axes, disease frameworks, cell-cell interactions, conceptual advances, or emerging cross-study topics.

Rules for review_extension_potential:
- High: the paper naturally opens a broad and timely review theme with clear relevance beyond this single study
- Medium: the paper supports a useful but somewhat narrower or more specialized review theme
- Low: the paper is mainly a discrete finding, technical advance, or narrowly scoped result with limited review-extension value

Rules for invited_review_angle:
- Write in English.
- Prefer 6 to 16 words.
- It should be broader than the original paper title, but still specific.
- It should be suitable as the central theme of a review article.
- It should reflect a mechanism, biological axis, conceptual framework, disease framework, or emerging topic.
- It should not simply repeat or lightly paraphrase the original title.
- It should not be a generic label such as "Neurodegeneration", "Brain function", or "Neural circuits".
- It should not be a keyword list.
- It should not be an overly narrow phrase tied only to one experiment, one dataset, or one isolated molecular detail unless that detail clearly represents a broader emerging theme.
- Prefer an angle that a scientific editor could realistically send in an invitation email.

Self-check before finalizing the angle:
- If your angle is too close to the original title, broaden it.
- If your angle is too broad to organize a coherent review, narrow it.
- If the paper is interesting but too narrow for a review, set review_extension_potential to Low and still provide the nearest reasonable angle.

Rules for angle_rationale_zh_cn:
- One concise sentence in Simplified Chinese.
- Explain why this angle is suitable for an invited review.
- Emphasize breadth, mechanism, framework value, or literature expandability.
- Do not exceed 45 Chinese characters if possible.

Journal: {journal}
Macro field: {macro_field}
Meso field: {meso_field}
Micro field: {micro_field}

Title:
{title}

Abstract:
{abstract}
""".strip()

    result = llm_generate(
        prompt,
        REVIEW_ANGLE_SCHEMA,
        audit_meta={
            "pmid": pmid,
            "stage": "review_angle",
            "title": title,
            "abstract": abstract,
        },
    )
    potential = normalize_review_extension_potential(result.get("review_extension_potential", "Medium"))
    angle = clean_space(result.get("invited_review_angle", ""))[:300]
    rationale = clean_space(result.get("angle_rationale_zh_cn", ""))[:300]
    return potential, angle, rationale


def translate_title_to_chinese(pmid: str, title: str) -> str:
    prompt = f"""
You are helping a scientific editor translate a biomedical paper title into Simplified Chinese.

Return JSON only.

Rules:
- Translate only the title.
- Use accurate academic Chinese.
- Keep gene/protein abbreviations, species names, and standard technical abbreviations when appropriate.
- Do not add explanation.
- Do not use quotation marks.

Title:
{title}
""".strip()

    result = llm_generate(
        prompt,
        TITLE_TRANSLATION_SCHEMA,
        audit_meta={
            "pmid": pmid,
            "stage": "title_translation",
            "title": title,
            "abstract": "",
        },
    )
    return clean_space(result.get("title_zh_cn", ""))[:500]


def suggest_field_hierarchy(pmid: str, title: str, abstract: str) -> Tuple[str, str, str]:
    prompt = f"""
You are helping a scientific editor classify a neuroscience-related paper for Neuroscience Bulletin.

Return JSON only.

Based on the Title and Abstract, provide:
1) macro_research_field: the broad discipline area
2) meso_research_field: the main subfield or systems-level topic
3) micro_research_field: the most specific mechanism, cell type, circuit, disease, method, or molecular focus

Rules:
- all three fields must be concise noun phrases
- use English
- each field should preferably be 2 to 10 words
- macro should be broadest, meso intermediate, micro most specific
- do not repeat the title
- avoid full sentences
- keep the hierarchy logically nested
- macro_research_field must be EXACTLY ONE label from the fixed list below
- do not invent a new macro label
- choose the label that best reflects the paper's central biological question rather than a peripheral method, disease context, or application
- if a paper uses a method but is not primarily about method development, do not choose "Neurotechnology, Imaging & Methods"
- if several labels seem relevant, choose the single best primary label

Fixed macro field list:
{NEURO_MACRO_FIELDS_TEXT}

Title: {title}

Abstract:
{abstract}
""".strip()

    result = llm_generate(
        prompt,
        FIELD_HIERARCHY_SCHEMA,
        audit_meta={
            "pmid": pmid,
            "stage": "field_hierarchy",
            "title": title,
            "abstract": abstract,
        },
    )
    return (
        clean_space(result.get("macro_research_field", ""))[:200],
        clean_space(result.get("meso_research_field", ""))[:200],
        clean_space(result.get("micro_research_field", ""))[:200],
    )


# ------------------------------------------------------------------
# Write Author sheet
# ------------------------------------------------------------------

def ensure_author_extra_columns(ws) -> None:
    template_row = 2 if ws.max_row >= 2 else 1
    for col, header in AUTHOR_EXTRA_HEADERS.items():
        current = clean_space(str(ws.cell(1, col).value or ""))
        if current != header:
            dst_header = ws.cell(1, col)
            # copy style from the last existing header cell (usually N1)
            donor_col = min(14, ws.max_column if ws.max_column >= 1 else 14)
            donor = ws.cell(1, donor_col)
            if donor.has_style:
                dst_header._style = copy(donor._style)
            if donor.font:
                dst_header.font = copy(donor.font)
            if donor.fill:
                dst_header.fill = copy(donor.fill)
            if donor.border:
                dst_header.border = copy(donor.border)
            if donor.alignment:
                dst_header.alignment = copy(donor.alignment)
            if donor.protection:
                dst_header.protection = copy(donor.protection)
            if donor.number_format:
                dst_header.number_format = donor.number_format
            dst_header.value = header

            # copy column style from the last existing data column (usually N)
            donor_data = ws.cell(template_row, donor_col)
            dst_data = ws.cell(template_row, col)
            if donor_data.has_style:
                dst_data._style = copy(donor_data._style)
            if donor_data.font:
                dst_data.font = copy(donor_data.font)
            if donor_data.fill:
                dst_data.fill = copy(donor_data.fill)
            if donor_data.border:
                dst_data.border = copy(donor_data.border)
            if donor_data.alignment:
                dst_data.alignment = copy(donor_data.alignment)
            if donor_data.protection:
                dst_data.protection = copy(donor_data.protection)
            if donor_data.number_format:
                dst_data.number_format = donor_data.number_format


def append_author_rows(
    target_xlsm: Path,
    sheet_name: str,
    rows_to_write: List[Tuple[ParsedArticle, str, str, str, str, str, str, str, str, str, dict]],
    invite_date: date,
) -> int:
    for idx, item in enumerate(rows_to_write, start=1):
        if len(item) != 11:
            raise ValueError(
                f"rows_to_write[{idx}] has {len(item)} items, expected 11 "
                "(article, topic, macro_field, meso_field, micro_field, title_cn, review_extension_potential, invited_review_angle, angle_rationale_zh_cn, editorial_bucket, author_identity)"
            )
    wb = load_workbook(target_xlsm, keep_vba=True)
    ws = wb[sheet_name]
    ensure_author_extra_columns(ws)

    written = 0
    template_row = 2 if ws.max_row >= 2 else 1

    for article, topic, macro_field, meso_field, micro_field, title_cn, review_extension_potential, invited_review_angle, angle_rationale_zh_cn, editorial_bucket, author_identity in rows_to_write:
        row = ws.max_row + 1

        # copy style from template row cell-by-cell
        for col in range(1, max(AUTHOR_EXTRA_HEADERS) + 1):
            src = ws.cell(template_row, col)
            dst = ws.cell(row, col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)

        ws.cell(row, 1, value=article.last_author_family_name)
        ws.cell(row, 2, value=article.last_author_email)
        ws.cell(row, 3, value=article.journal)
        ws.cell(row, 4, value=article.title)
        ws.cell(row, 5, value=article.last_author_full_name)
        ws.cell(row, 6, value=article.publication_date)
        ws.cell(row, 7, value=article.last_author_affiliation)
        ws.cell(row, 8, value=article.last_author_country)
        ws.cell(row, 9, value=article.pubmed_link)
        ws.cell(row, 10, value=topic)
        ws.cell(row, 11, value=article.first_author_full_name)
        ws.cell(row, 12, value=article.first_author_email)
        ws.cell(row, 13, value="No" if article.last_author_country.strip().lower() in CHINA_EQUIVALENTS.union({"china"}) else "Yes")
        ws.cell(row, 14, value=invite_date)
        ws.cell(row, 15, value=macro_field)
        ws.cell(row, 16, value=meso_field)
        ws.cell(row, 17, value=micro_field)
        ws.cell(row, 18, value=title_cn)
        ws.cell(row, 19, value=review_extension_potential)
        ws.cell(row, 20, value=invited_review_angle)
        ws.cell(row, 21, value=angle_rationale_zh_cn)
        ws.cell(row, 22, value=editorial_bucket)
        ws.cell(row, 23, value="")
        ws.cell(row, 24, value=getattr(article, "neuro_decision_source", ""))
        ws.cell(row, 25, value=article.pmid)

        style_written_row(ws, row, article, author_identity)

        written += 1

    apply_author_sheet_layout(ws)
    wb.save(target_xlsm)
    return written



def apply_author_sheet_layout(ws) -> None:
    # Wrap text for D, J, O, P, Q
    wrap_cols = [4, 10, 15, 16, 17, 18, 20, 21, 22]
    for col in wrap_cols:
        letter = ws.cell(1, col).column_letter
        ws.column_dimensions[letter].width = {
            4: 72,   # Title
            10: 42,  # Research field
            15: 30,  # Macro
            16: 30,  # Meso
            17: 30,  # Micro
            18: 42,  # Title Chinese
            19: 16,  # Review Potential
            20: 42,  # Invited Review Angle
            21: 42,  # Angle Rationale CN
            22: 22,  # Editorial Bucket
            23: 16,  # Manual Decision
            24: 18,  # Neuro Decision Source
        25: 12,  # PMID
            25: 12,  # PMID
        }.get(col, 24)

    # Improve readability for common columns
    fixed_widths = {
        1: 14,   # Family Name
        2: 28,   # Email
        3: 14,   # Journal
        5: 22,   # Full name
        6: 12,   # Date of Publication
        7: 26,   # Affiliation
        8: 14,   # Country
        9: 26,   # PubMed link
        11: 22,  # First author full name
        12: 26,  # First author email
        13: 10,  # Overseas
        14: 12,  # Date of Invitation
        22: 20,  # Editorial Bucket
        23: 16,  # Manual Decision
        24: 18,  # Neuro Decision Source
    }
    for col, width in fixed_widths.items():
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def style_written_row(ws, row: int, article, author_identity: dict) -> None:
    # D, J, O, P, Q wrap text
    for col in [4, 10, 15, 16, 17, 18, 20, 21, 22]:
        cell = ws.cell(row, col)
        base = copy(cell.alignment) if cell.alignment else Alignment()
        cell.alignment = Alignment(
            horizontal=base.horizontal,
            vertical="top",
            text_rotation=base.text_rotation or 0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=base.indent or 0,
        )

    # Give the row a readable height
    ws.row_dimensions[row].height = 42

    # Date formats
    ws.cell(row, 6).number_format = "yyyy/m/d"
    ws.cell(row, 14).number_format = "yyyy/m/d"

    # Country = China -> orange font for the whole row
    if article.last_author_country.strip().lower() in CHINA_EQUIVALENTS.union({"china"}):
        for col in range(1, max(AUTHOR_EXTRA_HEADERS) + 1):
            cell = ws.cell(row, col)
            cell.font = orange_font_like(cell.font)

    # Chinese-name = Yes -> light blue fill for the whole row
    if author_identity.get("last_author_is_chinese_name") == "Yes" or author_identity.get("first_author_is_chinese_name") == "Yes":
        for col in range(1, max(AUTHOR_EXTRA_HEADERS) + 1):
            ws.cell(row, col).fill = copy(LIGHT_BLUE_ROW_FILL)

    # Email mismatch -> only the email cells get red fill
    if author_identity.get("last_author_email_match") == "No" and clean_space(article.last_author_email):
        ws.cell(row, 2).fill = copy(RED_FLUORESCENT_FILL)
    if author_identity.get("first_author_email_match") == "No" and clean_space(article.first_author_email):
        ws.cell(row, 12).fill = copy(RED_FLUORESCENT_FILL)

def editorial_bucket_from_potential(potential: str) -> str:
    p = clean_space(potential).lower()
    if p == "high":
        return "Priority Invite"
    if p == "medium":
        return "Review"
    if p == "low":
        return "Low Priority Review"
    return "Review"


def write_run_summary(
    run_date: date,
    from_date: date,
    to_date: date,
    summary: dict,
    selected_rows: list[dict],
) -> Path:
    out = RUN_LOG_DIR / f"run-summary-{fmt_compact(run_date)}.md"
    lines = [
        f"# Review Invitation Agent Summary - {fmt_compact(run_date)}",
        "",
        f"- Date window: {fmt_slash(from_date)} -> {fmt_slash(to_date)}",
        f"- PubMed hits: {summary.get('hits', 0)}",
        f"- Parsed records: {summary.get('parsed', 0)}",
        f"- Written to Author sheet: {summary.get('written', 0)}",
        f"- Priority Invite: {summary.get('review_potential_high', 0)}",
        f"- Review: {summary.get('review_potential_medium', 0)}",
        f"- Low Priority Review: {summary.get('review_potential_low', 0)}",
        f"- Skipped no abstract: {summary.get('skipped_no_abstract', 0)}",
        f"- Skipped short abstract: {summary.get('skipped_short_abstract', 0)}",
        f"- Missing last author email: {summary.get('missing_last_author_email', 0)}",
        f"- Skipped duplicate: {summary.get('skipped_duplicate', 0)}",
        f"- Skipped non-neuro: {summary.get('skipped_non_neuro', 0)}",
        f"- Errors: {summary.get('errors', 0)}",
        "",
        "## Selected papers",
        "",
    ]

    if not selected_rows:
        lines.append("No papers written.")
    else:
        for row in selected_rows:
            lines.extend([
                f"### PMID {row['pmid']}",
                f"- Journal: {row['journal']}",
                f"- Title: {row['title']}",
                f"- Review Extension Potential: {row['review_extension_potential']}",
                f"- Editorial Bucket: {row['editorial_bucket']}",
                f"- Neuro Decision Source: {row.get('neuro_decision_source', '')}",
                f"- Invited Review Angle: {row['invited_review_angle']}",
                f"- Angle Rationale (CN): {row['angle_rationale_zh_cn']}",
                "",
            ])

    out.write_text("\n".join(lines), encoding="utf-8")
    return out


# ------------------------------------------------------------------
# Run log
# ------------------------------------------------------------------

def write_tsv_report(run_date: date, lines: List[str]) -> Path:
    out = RUN_LOG_DIR / f"run-{fmt_compact(run_date)}.tsv"
    out.write_text("\n".join(lines) if lines else "No records processed.", encoding="utf-8")
    return out


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main() -> int:
    ensure_dirs()
    validate_runtime_config()

    run_date = date.today()
    last_to = read_last_to_date(LOG_XLSX, LOG_SHEET)
    from_date = last_to + timedelta(days=1)
    to_date = run_date

    if from_date > to_date:
        print(f"No new window to process. Last To date = {fmt_slash(last_to)}")
        return 0

    processed_pmids = load_processed_pmids()
    existing_keys, existing_pmids_in_sheet = load_existing_author_keys(TARGET_XLSM, TARGET_SHEET)

    query = build_pubmed_query(from_date, to_date)
    print(f"Date window: {fmt_slash(from_date)} -> {fmt_slash(to_date)}")
    print(f"Query: {query}")

    summary = {
        "hits": 0,
        "parsed": 0,
        "skipped_no_abstract": 0,
        "skipped_short_abstract": 0,
        "missing_last_author_email": 0,
        "skipped_duplicate": 0,
        "skipped_non_neuro": 0,
        "written": 0,
        "review_potential_high": 0,
        "review_potential_medium": 0,
        "review_potential_low": 0,
        "errors": 0,
    }
    report_lines = []
    selected_rows = []

    pmids = esearch_pubmed(query)
    summary["hits"] = len(pmids)
    print(f"PubMed hits: {len(pmids)}")

    xml_text = efetch_pubmed_xml(pmids)
    xml_path = save_xml(xml_text, to_date)
    print(f"Saved XML: {xml_path}")

    records = parse_pubmed_xml(xml_text)
    summary["parsed"] = len(records)

    to_write: List[Tuple[ParsedArticle, str, str, str, str, str, str, str, str, str, dict]] = []

    total = len(records)
    for i, rec in enumerate(records, start=1):
        try:
            row_key = (rec.last_author_email.lower(), rec.journal.lower(), rec.title.lower())
            if rec.pmid in processed_pmids or rec.pmid in existing_pmids_in_sheet or row_key in existing_keys:
                summary["skipped_duplicate"] += 1
                print(f"[{i}/{total}] PMID {rec.pmid} | duplicate | skipped")
                report_lines.append(f"{rec.pmid}\tDUPLICATE\t{rec.title}")
                continue

            if not rec.abstract:
                summary["skipped_no_abstract"] += 1
                print(f"[{i}/{total}] PMID {rec.pmid} | no abstract | skipped")
                report_lines.append(f"{rec.pmid}	NO_ABSTRACT	{rec.title}")
                processed_pmids.add(rec.pmid)
                continue

            abstract_len = len(rec.abstract.strip())

            if PRINT_ABSTRACT_LENGTH:
                print(f"[{i}/{total}] PMID {rec.pmid} | abstract_len={abstract_len}")
                if abstract_len < ABSTRACT_SHORT_WARNING:
                    print(f"WARNING: PMID {rec.pmid} abstract looks short: {abstract_len} chars")

            if abstract_len < ABSTRACT_MIN_LENGTH:
                summary["skipped_short_abstract"] += 1
                print(f"[{i}/{total}] PMID {rec.pmid} | short abstract (<{ABSTRACT_MIN_LENGTH}) | skipped")
                report_lines.append(f"{rec.pmid}	SHORT_ABSTRACT	len={abstract_len}	{rec.title}")
                processed_pmids.add(rec.pmid)
                continue

            if not rec.last_author_email.strip():
                summary["missing_last_author_email"] += 1
                print(f"[{i}/{total}] PMID {rec.pmid} | missing last author email | continue")

            label, conf, reason, decision_source = decide_neuroscience_status(rec.pmid, rec.journal, rec.title, rec.abstract)
            rec.neuro_decision_source = decision_source
            if label != "Yes":
                summary["skipped_non_neuro"] += 1
                print(f"[{i}/{total}] PMID {rec.pmid} | neuro=No ({conf:.2f}) | skipped")
                report_lines.append(f"{rec.pmid}\tNON_NEURO\t{conf:.2f}\t{reason}\t{rec.title}")
                processed_pmids.add(rec.pmid)
                continue

            author_identity = assess_author_identity(
                rec.pmid,
                rec.last_author_full_name,
                rec.last_author_email,
                rec.first_author_full_name,
                rec.first_author_email,
            )
            topic = suggest_topic(rec.pmid, rec.title, rec.abstract)
            macro_field, meso_field, micro_field = suggest_field_hierarchy(rec.pmid, rec.title, rec.abstract)
            title_cn = translate_title_to_chinese(rec.pmid, rec.title)
            review_extension_potential, invited_review_angle, angle_rationale_zh_cn = generate_review_angle_package(
                rec.pmid, rec.title, rec.abstract, rec.journal, macro_field, meso_field, micro_field
            )
            editorial_bucket = editorial_bucket_from_potential(review_extension_potential)

            if review_extension_potential == "High":
                summary["review_potential_high"] += 1
            elif review_extension_potential == "Medium":
                summary["review_potential_medium"] += 1
            else:
                summary["review_potential_low"] += 1

            to_write.append((rec, topic, macro_field, meso_field, micro_field, title_cn, review_extension_potential, invited_review_angle, angle_rationale_zh_cn, editorial_bucket, author_identity))
            processed_pmids.add(rec.pmid)
            existing_keys.add(row_key)
            existing_pmids_in_sheet.add(rec.pmid)
            selected_rows.append({
                "pmid": rec.pmid,
                "journal": rec.journal,
                "title": rec.title,
                "review_extension_potential": review_extension_potential,
                "editorial_bucket": editorial_bucket,
                "invited_review_angle": invited_review_angle,
                "angle_rationale_zh_cn": angle_rationale_zh_cn,
                "neuro_decision_source": rec.neuro_decision_source,
            })

            print(
                f"[{i}/{total}] PMID {rec.pmid} | neuro=Yes ({conf:.2f}) | "
                f"topic={topic} | macro={macro_field} | meso={meso_field} | micro={micro_field} | "
                f"potential={review_extension_potential} | bucket={editorial_bucket} | angle={invited_review_angle}"
            )
            report_lines.append(
                f"{rec.pmid}\tWRITE\t{conf:.2f}\t{topic}\t{macro_field}\t{meso_field}\t{micro_field}\t{title_cn}\t{review_extension_potential}\t{editorial_bucket}\t{invited_review_angle}\t{angle_rationale_zh_cn}\t{rec.title}"
            )

        except Exception as e:
            summary["errors"] += 1
            print(f"[{i}/{total}] PMID {rec.pmid} | ERROR | {e}")
            report_lines.append(f"{rec.pmid}\tERROR\t{e}\t{rec.title}")

    if to_write:
        summary["written"] = append_author_rows(TARGET_XLSM, TARGET_SHEET, to_write, invite_date=run_date)

    report_path = write_tsv_report(run_date, report_lines)
    summary_path = write_run_summary(run_date, from_date, to_date, summary, selected_rows)

    # append From / To only after the whole run succeeds
    append_from_to_log(LOG_XLSX, LOG_SHEET, from_date, to_date)
    save_processed_pmids(processed_pmids)

    print("\nDone.")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"XML saved to: {xml_path}")
    print(f"TSV report saved to: {report_path}")
    print(f"Run summary saved to: {summary_path}")
    if LLM_AUDIT_ENABLED:
        print(f"LLM audit dir: {llm_audit_dir(run_date)}")
    return 0


def run_with_lock() -> int:
    acquire_run_lock(RUN_LOCK_FILE)
    try:
        return main()
    finally:
        release_run_lock(RUN_LOCK_FILE)

if __name__ == "__main__":
    raise SystemExit(run_with_lock())

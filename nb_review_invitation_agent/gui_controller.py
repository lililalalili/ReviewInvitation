from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from urllib.parse import quote_plus

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .workbook import get_header_map, load_workbook_preserve_vba

UNBATCHED = "Unbatched"
EDITABLE_FIELDS = ["Overseas", "Manual Decision", "Research field"]
DISPLAY_FIELDS = [
    "Full Name of the Last Author",
    "Affiliation of the Last Author",
    "Country",
    "Overseas",
    "Last Author Web",
    "Email of the Last Author",
    "First Author Full name",
    "First Author Email",
    "Journal",
    "Title",
    "Title (CN)",
    "Pubmed Link",
    "Research field",
    "Macro Research Field",
    "Meso Research Field",
    "Micro Research Field",
    "Review Extension Potential",
    "Invited Review Angle",
    "Angle Rationale (CN)",
    "Editorial Bucket",
    "Manual Decision",
    "Neuro Decision Source",
    "Date of Invitaion",
]


@dataclass
class RowView:
    row_number: int
    values: dict[str, str]


class WorkbookReviewController:
    def __init__(self, workbook: Workbook, worksheet: Worksheet) -> None:
        self.workbook = workbook
        self.worksheet = worksheet
        self.header_map = get_header_map(worksheet)
        self.batch_rows = self._build_batch_rows()
        self.batch_ids = list(self.batch_rows.keys())
        self.selected_batch = self._default_batch()
        self.current_index = 0
        self.save_path: str | Path | None = None
        self._ensure_current_batch_index()

    @classmethod
    def from_xlsm_path(cls, path: str | Path) -> "WorkbookReviewController":
        wb = load_workbook_preserve_vba(path)
        ws = wb["Author"] if "Author" in wb.sheetnames else wb.active
        instance = cls(wb, ws)
        instance.save_path = path
        return instance

    def _default_batch(self) -> str:
        last_non_empty = None
        batch_col = self.header_map.get("Batch ID")
        if batch_col is not None:
            for row in range(2, self.worksheet.max_row + 1):
                value = self.worksheet.cell(row=row, column=batch_col).value
                text = str(value).strip() if value is not None else ""
                if text:
                    last_non_empty = text
        if last_non_empty:
            return last_non_empty
        return UNBATCHED

    def _build_batch_rows(self) -> dict[str, list[int]]:
        rows: dict[str, list[int]] = {}
        batch_col = self.header_map.get("Batch ID")
        for row in range(2, self.worksheet.max_row + 1):
            batch_value = ""
            if batch_col is not None:
                raw = self.worksheet.cell(row=row, column=batch_col).value
                batch_value = str(raw).strip() if raw is not None else ""
            batch_id = batch_value or UNBATCHED
            rows.setdefault(batch_id, []).append(row)
        if UNBATCHED not in rows:
            rows[UNBATCHED] = []
        return rows

    def _ensure_current_batch_index(self) -> None:
        if self.selected_batch not in self.batch_rows:
            self.selected_batch = UNBATCHED
        max_idx = max(len(self.batch_rows[self.selected_batch]) - 1, 0)
        self.current_index = min(self.current_index, max_idx)

    def set_batch(self, batch_id: str) -> None:
        self.selected_batch = batch_id if batch_id in self.batch_rows else UNBATCHED
        self.current_index = 0
        self._ensure_current_batch_index()

    def _current_row_number(self) -> int | None:
        rows = self.batch_rows.get(self.selected_batch, [])
        if not rows:
            return None
        return rows[self.current_index]

    def get_current_row(self) -> RowView:
        row_number = self._current_row_number()
        values = {header: "" for header in DISPLAY_FIELDS}
        if row_number is None:
            return RowView(row_number=0, values=values)
        for header in DISPLAY_FIELDS:
            col = self.header_map.get(header)
            if col is None:
                continue
            cell_value = self.worksheet.cell(row=row_number, column=col).value
            values[header] = "" if cell_value is None else str(cell_value)
        return RowView(row_number=row_number, values=values)

    def update_current_row(self, updates: dict[str, str]) -> None:
        row_number = self._current_row_number()
        if row_number is None:
            return
        for field, value in updates.items():
            if field not in EDITABLE_FIELDS:
                continue
            col = self.header_map.get(field)
            if col is None:
                continue
            self.worksheet.cell(row=row_number, column=col, value=value)

    def navigate_next(self) -> None:
        rows = self.batch_rows.get(self.selected_batch, [])
        if rows and self.current_index < len(rows) - 1:
            self.current_index += 1

    def navigate_previous(self) -> None:
        if self.current_index > 0:
            self.current_index -= 1

    def navigate_first(self) -> None:
        self.current_index = 0

    def navigate_last(self) -> None:
        rows = self.batch_rows.get(self.selected_batch, [])
        if rows:
            self.current_index = len(rows) - 1

    def save_workbook(self, path: str | Path) -> None:
        try:
            self.workbook.save(path)
        except PermissionError as exc:
            raise RuntimeError("Failed to save workbook. Please close NB_Author_2026.xlsm in Excel and try again.") from exc
        except OSError as exc:
            if getattr(exc, "errno", None) == 13:
                raise RuntimeError("Failed to save workbook. Please close NB_Author_2026.xlsm in Excel and try again.") from exc
            raise

    def set_row_value(self, row_number: int, field: str, value: str) -> None:
        col = self.header_map.get(field)
        if col is None:
            return
        self.worksheet.cell(row=row_number, column=col, value=value)

    def get_selected_batch_rows(self) -> list[RowView]:
        rows = self.batch_rows.get(self.selected_batch, [])
        output: list[RowView] = []
        for row_number in rows:
            values = {header: "" for header in DISPLAY_FIELDS}
            for header in DISPLAY_FIELDS:
                col = self.header_map.get(header)
                if col is None:
                    continue
                raw = self.worksheet.cell(row=row_number, column=col).value
                values[header] = "" if raw is None else str(raw)
            output.append(RowView(row_number=row_number, values=values))
        return output


def build_email_search_url(email: str) -> str:
    return f"https://www.bing.com/search?q={quote_plus(email)}"


def default_open_url(url: str) -> None:
    import webbrowser

    webbrowser.open(url)


def invitation_placeholder() -> str:
    return "Invitation sending is implemented in Task 05"

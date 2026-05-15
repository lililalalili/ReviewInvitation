from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_COLUMNS = [
    "Affiliation of the First Author",
    "Last Author Research",
    "Last Author Web",
    "Batch ID",
    "Author Enrichment Status",
    "Author Enrichment Evidence",
    "Invitation Status",
    "Invitation Error",
]

DATE_OF_INVITAION = "Date of Invitaion"
PUBMED_LINK = "Pubmed Link"
PMID_HEADER = "PMID"


def load_workbook_preserve_vba(path: str | Path):
    return openpyxl.load_workbook(path, keep_vba=True)


def get_header_map(ws: Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        if header and header not in header_map:
            header_map[header] = col_idx
    return header_map


def ensure_columns(ws: Worksheet, required_columns: list[str]) -> dict[str, int]:
    header_map = get_header_map(ws)
    for header in required_columns:
        if header not in header_map:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value=header)
            header_map[header] = col_idx
    return header_map


def generate_batch_id(now: datetime | None = None) -> str:
    now = now or datetime.now()
    return now.strftime("%Y%m%d_%H%M%S")


def build_pubmed_article_link(pmid: str | int | None) -> str:
    if pmid is None:
        return ""
    pmid_text = str(pmid).strip()
    if not pmid_text:
        return ""
    return f"https://pubmed.ncbi.nlm.nih.gov/{pmid_text}/"


def prepare_new_author_row(row_data: dict[str, Any], batch_id: str) -> dict[str, Any]:
    prepared = dict(row_data)
    prepared["Batch ID"] = batch_id
    prepared[DATE_OF_INVITAION] = ""

    if not prepared.get(PUBMED_LINK):
        prepared[PUBMED_LINK] = build_pubmed_article_link(prepared.get(PMID_HEADER))

    return prepared


def append_author_row(ws: Worksheet, row_data: dict[str, Any], batch_id: str) -> int:
    headers = ensure_columns(ws, REQUIRED_COLUMNS)
    headers = ensure_columns(ws, [DATE_OF_INVITAION, PUBMED_LINK, PMID_HEADER])

    prepared = prepare_new_author_row(row_data, batch_id)
    new_row_num = ws.max_row + 1

    for header, col_idx in headers.items():
        ws.cell(row=new_row_num, column=col_idx, value=prepared.get(header, ""))

    return new_row_num


def mark_invited(ws: Worksheet, row_number: int, invite_date: str) -> None:
    headers = ensure_columns(ws, [DATE_OF_INVITAION])
    ws.cell(row=row_number, column=headers[DATE_OF_INVITAION], value=invite_date)

from datetime import datetime

from openpyxl import Workbook

from nb_review_invitation_agent.workbook import (
    DATE_OF_INVITAION,
    PUBMED_LINK,
    append_author_row,
    build_pubmed_article_link,
    ensure_columns,
    generate_batch_id,
    get_header_map,
    load_workbook_preserve_vba,
    mark_invited,
)


def make_sheet_with_headers(headers: list[str]):
    wb = Workbook()
    ws = wb.active
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    return wb, ws


def test_ensure_columns_appends_missing_once():
    _, ws = make_sheet_with_headers(["PMID", "Date of Invitaion"])
    required = ["Batch ID", "Invitation Status", "Invitation Error"]

    ensure_columns(ws, required)
    ensure_columns(ws, required)

    headers = [c.value for c in ws[1] if c.value is not None]
    assert headers.count("Batch ID") == 1
    assert headers.count("Invitation Status") == 1
    assert headers.count("Invitation Error") == 1


def test_append_author_row_keeps_invitation_date_blank_and_assigns_batch():
    _, ws = make_sheet_with_headers(
        ["PMID", PUBMED_LINK, DATE_OF_INVITAION, "Manual Decision", "Date of Publication"]
    )
    batch_id = generate_batch_id(datetime(2026, 5, 15, 14, 30, 12))
    row_data = {
        "PMID": "12345678",
        "Date of Publication": "2026-05-15",
        "Manual Decision": "",
        "Affiliation of the First Author": "Neuroscience Institute",
    }

    row_number = append_author_row(ws, row_data, batch_id)
    header_map = get_header_map(ws)

    assert ws.cell(row=row_number, column=header_map[DATE_OF_INVITAION]).value == ""
    assert ws.cell(row=row_number, column=header_map["Batch ID"]).value == batch_id
    assert ws.cell(row=row_number, column=header_map["Affiliation of the First Author"]).value == "Neuroscience Institute"
    assert ws.cell(row=row_number, column=header_map[PUBMED_LINK]).value == "https://pubmed.ncbi.nlm.nih.gov/12345678/"


def test_mark_invited_only_writes_when_explicitly_called():
    _, ws = make_sheet_with_headers(["PMID", DATE_OF_INVITAION])
    row_number = append_author_row(ws, {"PMID": "12345678"}, "20260515_143012")
    header_map = get_header_map(ws)

    assert ws.cell(row=row_number, column=header_map[DATE_OF_INVITAION]).value == ""

    mark_invited(ws, row_number, "2026-05-15")
    assert ws.cell(row=row_number, column=header_map[DATE_OF_INVITAION]).value == "2026-05-15"


def test_batch_id_same_for_rows_in_one_batch():
    _, ws = make_sheet_with_headers(["PMID", DATE_OF_INVITAION])
    batch_id = generate_batch_id(datetime(2026, 5, 15, 14, 30, 12))

    row1 = append_author_row(ws, {"PMID": "111"}, batch_id)
    row2 = append_author_row(ws, {"PMID": "222"}, batch_id)
    header_map = get_header_map(ws)

    assert ws.cell(row=row1, column=header_map["Batch ID"]).value == batch_id
    assert ws.cell(row=row2, column=header_map["Batch ID"]).value == batch_id


def test_build_pubmed_article_link_pmid_format():
    assert build_pubmed_article_link("12345678") == "https://pubmed.ncbi.nlm.nih.gov/12345678/"


def test_load_workbook_preserve_vba_uses_keep_vba(monkeypatch):
    called = {}

    def fake_load_workbook(path, keep_vba=False):
        called["path"] = path
        called["keep_vba"] = keep_vba
        return "ok"

    monkeypatch.setattr("nb_review_invitation_agent.workbook.openpyxl.load_workbook", fake_load_workbook)

    result = load_workbook_preserve_vba("sample.xlsm")

    assert result == "ok"
    assert called["path"] == "sample.xlsm"
    assert called["keep_vba"] is True

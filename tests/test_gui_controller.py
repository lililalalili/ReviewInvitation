import pytest
from openpyxl import Workbook

from nb_review_invitation_agent.gui_controller import (
    DISPLAY_FIELDS,
    UNBATCHED,
    WorkbookReviewController,
    build_email_search_url,
)


def make_controller():
    wb = Workbook()
    ws = wb.active
    ws.title = "Author"
    headers = DISPLAY_FIELDS + ["Batch ID"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)

    rows = [
        {"Batch ID": "", "Full Name of the Last Author": "A", "Overseas": "No", "Manual Decision": "No", "Research field": "x", "Date of Invitaion": ""},
        {"Batch ID": "B1", "Full Name of the Last Author": "B", "Overseas": "Yes", "Manual Decision": "Review", "Research field": "y", "Date of Invitaion": ""},
        {"Batch ID": "", "Full Name of the Last Author": "C", "Overseas": "No", "Manual Decision": "Insight", "Research field": "z", "Date of Invitaion": ""},
        {"Batch ID": "B2", "Full Name of the Last Author": "D", "Overseas": "Yes", "Manual Decision": "No", "Research field": "k", "Date of Invitaion": ""},
    ]
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    return WorkbookReviewController(wb, ws)


def test_batch_discovery_and_default_latest_batch():
    c = make_controller()
    assert "B1" in c.batch_ids and "B2" in c.batch_ids
    assert UNBATCHED in c.batch_ids
    assert c.selected_batch == "B2"


def test_select_batch_returns_first_row():
    c = make_controller()
    c.set_batch(UNBATCHED)
    assert c.get_current_row().values["Full Name of the Last Author"] == "A"


def test_navigation_within_batch():
    c = make_controller()
    c.set_batch(UNBATCHED)
    assert c.get_current_row().values["Full Name of the Last Author"] == "A"
    c.navigate_next()
    assert c.get_current_row().values["Full Name of the Last Author"] == "C"
    c.navigate_next()
    assert c.get_current_row().values["Full Name of the Last Author"] == "C"
    c.navigate_previous()
    assert c.get_current_row().values["Full Name of the Last Author"] == "A"
    c.navigate_last()
    assert c.get_current_row().values["Full Name of the Last Author"] == "C"
    c.navigate_first()
    assert c.get_current_row().values["Full Name of the Last Author"] == "A"


def test_editable_fields_update_and_readonly_not_written(tmp_path):
    c = make_controller()
    c.set_batch("B2")
    before_title = c.get_current_row().values["Title"]
    c.update_current_row({"Overseas": "No", "Manual Decision": "Review", "Research field": "abc", "Title": "hacked"})
    row = c.get_current_row().values
    assert row["Overseas"] == "No"
    assert row["Manual Decision"] == "Review"
    assert row["Research field"] == "abc"
    assert row["Title"] == before_title
    p = tmp_path / "fake.xlsm"
    c.save_workbook(p)


def test_link_builders():
    assert build_email_search_url("a+b@example.edu") == "https://www.bing.com/search?q=a%2Bb%40example.edu"


def test_save_workbook_permission_error_message(monkeypatch, tmp_path):
    c = make_controller()

    def _raise(_):
        raise PermissionError("[Errno 13] Permission denied")

    monkeypatch.setattr(c.workbook, "save", _raise)
    with pytest.raises(RuntimeError, match="Please close NB_Author_2026.xlsm"):
        c.save_workbook(tmp_path / "x.xlsm")
